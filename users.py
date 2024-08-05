import logging
import requests
import openpyxl

# Configure logging
logging.basicConfig(level=logging.INFO)

# Replace with your actual API key and Discourse domain
API_KEY = " "
DISCOURSE_DOMAIN = " "

# API endpoints
CREATE_USER_ENDPOINT = f"{DISCOURSE_DOMAIN}/users.json"
ACTIVATE_USER_ENDPOINT = f"{DISCOURSE_DOMAIN}/admin/users/{{}}/activate.json"

def handle_error(response):
    if response.status_code != 200:
        print(f"Error: {response.json().get('errors', response.text)}")

def create_user(ws, row, username, email, name, password):
    payload = {
        "api_key": API_KEY,
        "api_username": "easdevhub",  # Replace with your admin username
        "email": email,
        "username": username,
        "name": name,
        "password": password,
        "active": True,
        "approved": True,
        "suppress_welcome_message": True  # Suppress the welcome message
    }

    try:
        response = requests.post(CREATE_USER_ENDPOINT, json=payload)
        response.raise_for_status()  # Raise an exception for non-200 status codes
        response_data = response.json()
        
        if response_data.get('success'):
            user_id = response_data.get('user_id')
            ws.cell(row=row, column=5, value="User Created")  # Update User Status column
            ws.cell(row=row, column=6, value=str(response_data))  # Update API Response column
            ws.cell(row=row, column=7, value=user_id)  # Save the user ID in column G
            logging.info(f"User created successfully: {username} with ID {user_id}")
            return user_id
        else:
            handle_error(response)
            ws.cell(row=row, column=5, value="User Creation Failed")  # Update User Status column
            ws.cell(row=row, column=6, value=str(response_data))  # Update API Response column
            logging.error(f"Error creating user {username}: {response_data}")
            return None
            
    except requests.exceptions.RequestException as e:
        ws.cell(row=row, column=5, value="User Creation Failed")  # Update User Status column
        ws.cell(row=row, column=6, value=str(e))  # Update API Response column
        logging.error(f"Error creating user {username}: {e}")
        return None

def activate_user(ws, row, user_id):
    if user_id is None:
        ws.cell(row=row, column=8, value="Failed to create user")
        return

    url = ACTIVATE_USER_ENDPOINT.format(user_id)
    payload = {
        "api_key": API_KEY,
        "api_username": "easdevhub"  # Replace with your admin username
    }
    
    try:
        response = requests.put(url, json=payload)
        response.raise_for_status()  # Raise an exception for non-200 status codes
        response_data = response.json()
        
        if response_data.get('success'):
            ws.cell(row=row, column=8, value="Activated")
            logging.info(f"User with ID {user_id} activated successfully")
        else:
            handle_error(response)
            ws.cell(row=row, column=8, value="Activation failed")
            ws.cell(row=row, column=6, value=str(response_data))
            logging.error(f"Error activating user {user_id}: {response_data}")
            
    except requests.exceptions.RequestException as e:
        ws.cell(row=row, column=8, value="Activation failed")
        ws.cell(row=row, column=6, value=str(e))
        logging.error(f"Error activating user {user_id}: {e}")

def read_user_data(filename):
    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook.active

    # Add headers for status and API response
    worksheet.cell(row=1, column=5, value="User Status")
    worksheet.cell(row=1, column=6, value="API Response")
    worksheet.cell(row=1, column=7, value="User ID")
    worksheet.cell(row=1, column=8, value="Activation Status")

    for row, data in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        username, email, name, password = data[:4]  # Capture only the first four values
        if not password:
            ws.cell(row=row, column=5, value="No Password Provided")
            continue
        
        user_id = create_user(worksheet, row, username, email, name, password)
        activate_user(worksheet, row, user_id)

    workbook.save(filename)

# Example usage: Replace 'users.xlsx' with your actual Excel file path
read_user_data("users.xlsx")

print("User creation and activation process complete.")

